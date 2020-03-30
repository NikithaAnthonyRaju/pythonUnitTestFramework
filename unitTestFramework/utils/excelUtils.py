import openpyxl
import os

class readExcelData():

    def __init__(self):
        # self.project_root = os.path.dirname(os.path.dirname(__file__))
        # self.output_path = os.path.join(self.project_root, 'data')
        # self.file1 = self.output_path + "\\testData.xlsx"
        self.file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data\\testData.xlsx')
        self.workbook = openpyxl.load_workbook(self.file)
        self.workSheet = self.workbook.active

    def getRowCount(self):
        print('Row Count -- ' , self.workSheet.max_row)
        return (self.workSheet.max_row)

    def getColumnCount(self):
        print('Column Count -- ' , self.workSheet.max_column)
        return (self.workSheet.max_column)

    def readDate(self):
        for row in self.workSheet.iter_rows(min_row=2, max_row=self.getRowCount(), min_col=2, max_col=self.getColumnCount()):
            for cell in row:
                print(cell.value)

    def readDateAsList(self, data):
        excelData = []
        index = 0
        print("Path --- ", self.file)
        for row in self.workSheet.iter_rows(min_row=2, min_col=1, max_row=self.getRowCount(), max_col=1, values_only=True):
            for rowCell in row:
                #print("data--", data)
                #print("rowcell---", rowCell)
                if data == rowCell:
                    for col in self.workSheet.iter_cols(min_row=data+1, min_col=2, max_row=data+1,max_col=self.getColumnCount(), values_only=True):
                        for cell in col:
                            #print(cell)
                            excelData.insert(index, cell)
                            index = index + 1;
        return excelData

    def readDateAsDict(self, data):
       try:
           list = []
           excelData = {}
           index = 0
           for row in self.workSheet.iter_rows(min_row=2, min_col=1, max_row=self.getRowCount(), max_col=1,
                                               values_only=True):
               for rowCell in row:
                   if data == rowCell:
                       for row in self.workSheet.iter_rows(min_row=1, min_col=2, max_row=1,
                                                           max_col=self.getColumnCount(), values_only=True):
                           for rowCell in row:
                               list.append(rowCell)
                           for col in self.workSheet.iter_cols(min_row=data + 1, min_col=2, max_row=data + 1,
                                                               max_col=self.getColumnCount(), values_only=True):
                               for cell in col:
                                   print(cell)
                                   excelData[list[index]] = cell
                                   index = index + 1;
           return excelData
       except:
           print("Data not available")